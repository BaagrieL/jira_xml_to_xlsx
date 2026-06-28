import sys

import argparse

import html

import os

import glob

import xml.etree.ElementTree as ET

from openpyxl import Workbook

from openpyxl.styles import PatternFill, Font, Alignment

from datetime import datetime


ALIAS_PROPRIEDADES = {
    "key": "issue",
    "type": "tipo",
    "summary": "resumo",
    "description": "descrição",
    "created": "dt_criação",
    "updated": "última atualização",
    "reporter": "relator",
    "assignee": "responsável",
    "priority": "prioridade",
    "due": "prazo",
}

# Paleta de cores para tipos de issue
COLOR_PALETTE_TYPES = {
    "bug": "F3645A",
    "tarefa": "5C96EE",
    "história": "6A9B22",
    "orçamento": "6A9B22",
}

COLOR_PALETTE_PRIORITY = {
    "lowest": "11308b",
    "low": "1272ab",
    "medium": "b99900",
    "high": "830B0B",
    "highest": "c70000"
}

COLOR_PALETTE_STATUS = {
    "tarefas pendentes": "404040",
    "em andamento": "4688ec",
    "concluído": "388600",
    "rejeitado": "c70000"
}


def normalizar_nome(nome):
    if nome is None:
        return ""
    return " ".join(str(nome).split()).lower()


def obter_valor_tag(item, prop):
    tag = item.find(f".//{prop}")
    if tag is not None and tag.text:
        return html.unescape(tag.text.strip())
    return None


def obter_valor_customfield(item, prop):
    nome_alvo = normalizar_nome(prop)
    if not nome_alvo:
        return None

    for customfield in item.findall(".//customfield"):
        nome_tag = customfield.find("customfieldname")
        if nome_tag is None or not nome_tag.text:
            continue
        nome_campo = normalizar_nome(nome_tag.text)
        if nome_campo != nome_alvo:
            continue

        valores = []
        valor_tags = customfield.findall(".//customfieldvalue")
        for valor_tag in valor_tags:
            texto = (valor_tag.text or "").strip()
            if texto:
                valores.append(html.unescape(texto))
            elif valor_tag.get("key"):
                valores.append(valor_tag.get("key"))

        if valores:
            return "; ".join(valores)

        if valor_tags:
            return ""

        break

    return None


def aplicar_formatacao_tipo(cell, tipo_value):
    """
    Aplica formatação de cor ao texto baseado no tipo de issue.
    Se o tipo estiver na paleta, aplica a cor correspondente.
    Se não estiver, mantém a formatação padrão.
    """
    if tipo_value:
        tipo_normalizado = tipo_value.lower().strip()
        cor = COLOR_PALETTE_TYPES.get(tipo_normalizado)
        
        if cor:
            cell.font = Font(color=cor, bold=True)

def aplicar_formatacao_priority(cell, priority_value):
    if priority_value:
        priority_normalizado = priority_value.lower().strip().replace(" (migrated)", "")
        cor = COLOR_PALETTE_PRIORITY.get(priority_normalizado)
        if cor:
            cell.font = Font(color=cor, bold=False)

def aplicar_formatacao_status(cell, status_value):
    if status_value:
        status_normalizado = status_value.lower().strip()
        cor = COLOR_PALETTE_STATUS.get(status_normalizado)
        if cor:
            cell.font = Font(color=cor, bold=False)

# rodar com:
# py criar_csv_issues.py "caminho/do aquivo.xml propriedade1/propriedade2/.../propriedadeN"
def extrair_propriedades(xml_path, propriedades):
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f"[ERRO] Ao ler o XML: {e}")
        sys.exit(1)
    except FileNotFoundError:
        print(f"[ERRO] Arquivo '{xml_path}' não encontrado.")
        sys.exit(1)

    itens = root.findall(".//item")
    resultados = []
    propriedades_encontradas = set()

    for item in itens:
        registro = {}
        for prop in propriedades:
            valor = obter_valor_tag(item, prop)
            if valor is None:
                valor = obter_valor_customfield(item, prop)

            if valor is not None:
                propriedades_encontradas.add(prop)
                registro[prop] = valor
            else:
                registro[prop] = ""
        resultados.append(registro)

    propriedades_faltantes = set(propriedades) - propriedades_encontradas
    return resultados, propriedades_faltantes


# TODO: decidir como deve ser a saída para pastas (talvez uma nova pasta com n arquivos.txt [n = quantidade de arquivos xml])
# TODO: adaptar para ler pasta inteira de XMLs e gerar o arquivo .txt final
def salvar_txt(resultados, propriedades, nome_saida_txt):
    with open(nome_saida_txt, "w", encoding="utf-8") as f:
        for item in resultados:
            for prop in propriedades:
                f.write(f"{prop}: {item.get(prop, '')}\n")
            f.write("\n")
    print(f"Arquivo TXT gerado: {nome_saida_txt}")


def salvar_csv_ou_pasta(xml_path, propriedades, nome_saida_csv):
    nome_saida_xlsx = nome_saida_csv.replace(".csv", ".xlsx")
    wb = Workbook()

    if os.path.isdir(xml_path):
        arquivos_xml = glob.glob(os.path.join(xml_path, "*.xml"))
        if not arquivos_xml:
            print("Nenhum arquivo XML encontrado na pasta informada.")
            return

        for xml_file in arquivos_xml:
            nome_aba = os.path.splitext(os.path.basename(xml_file))[0][:31]
            ws = wb.create_sheet(title=nome_aba)
            ws.row_dimensions[1].height = 25
            print(f"Lendo {xml_file} -> aba '{nome_aba}'")

            resultados, _ = extrair_propriedades(xml_file, propriedades)
            _preparar_planilha(ws, propriedades, resultados)

        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
    else:
        ws = wb.active
        ws.title = os.path.splitext(os.path.basename(xml_path))[0][:31]
        ws.row_dimensions[1].height = 25
        resultados, _ = extrair_propriedades(xml_path, propriedades)
        _preparar_planilha(ws, propriedades, resultados)

    wb.save(nome_saida_xlsx)
    print(f"Arquivo XLSX gerado com sucesso: {nome_saida_xlsx}")


def _preparar_planilha(ws, propriedades, resultados):
    for col_idx, prop in enumerate(propriedades, 1):
        prop_alias = ALIAS_PROPRIEDADES.get(prop, prop)
        cell = ws.cell(row=1, column=col_idx, value=prop_alias)
        cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if prop in {"summary", "description"}:
            ws.column_dimensions[cell.column_letter].width = 105.7
            continue
        ws.column_dimensions[cell.column_letter].width = max(len(prop_alias) + 5, 15)

    for row_idx, item in enumerate(resultados, 2):
        for col_idx, prop in enumerate(propriedades, 1):

            if prop in {"created", "updated", "due"}:
                valor_formatado = formatar_data(item.get(prop, ""))
                cell = ws.cell(row=row_idx, column=col_idx, value=valor_formatado)
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=item.get(prop, ""))

            if prop in {"summary", "description"}:
                cell.alignment = Alignment(wrap_text=True, vertical="bottom")
            
            if prop == "type":
                aplicar_formatacao_tipo(cell, item.get("type", ""))
            if prop == "priority":
                aplicar_formatacao_priority(cell, item.get("priority", ""))
            if prop == "status":
                aplicar_formatacao_status(cell, item.get("status", ""))


def formatar_data(data_str):
    try:
        data_obj = datetime.strptime(data_str, "%a, %d %b %Y %H:%M:%S %z")
        return data_obj.strftime("%d/%m/%Y")
    except ValueError:
        return data_str


def main():
    parser = argparse.ArgumentParser(
        description="Extrai propriedades de um XML e gera CSV/TXT"
    )

    parser.add_argument("xml_path", help="Caminho do arquivo XML")

    parser.add_argument(
        "propriedades",
        nargs="?",
        help="Lista de propriedades separadas por /"
    )

    parser.add_argument(
        "-d",
        "--default",
        action="store_true",
        help="Retorna as propriedades padrão: \n type, \n key, \n summary, \n due, \n created, \n assignee, \n status, \n priority \n"
    )

    args = parser.parse_args()

    xml_path = args.xml_path

    if args.default:
        print("[INFO] Modo default ativado")
        propriedades = ["type", "key", "summary", "due", "created", "assignee", "status", "priority"]
    else:
        propriedades = args.propriedades.split("/")

    resultados = []
    faltantes = set()
    xml_path_base = os.path.splitext(xml_path)[0]

    if os.path.isfile(xml_path):
        resultados, faltantes = extrair_propriedades(xml_path, propriedades)

    if resultados:
        print("[OK] As seguintes propriedades foram encontradas:")
        props_encontradas = set(resultados[0].keys())
        for prop in propriedades:
            if prop in props_encontradas:
                print(f"        - {prop}")
    if faltantes:
        print("[AVISO] As seguintes propriedades não foram encontradas:")
        for prop in faltantes:
            print(f"        - {prop}")

    if xml_path_base:
        nome_saida_csv = xml_path_base
        nome_saida_txt = xml_path_base
    else:
        nome_saida_csv = "saida_" + "_".join(propriedades)
        nome_saida_txt = nome_saida_csv

    if os.path.isfile(xml_path):
        nome_saida_csv += ".csv"
        nome_saida_txt += ".txt"

    if os.path.isdir(xml_path):
        nome_saida_csv += os.path.basename(os.path.normpath(xml_path)) + ".csv"
        nome_saida_txt += os.path.basename(os.path.normpath(xml_path)) + ".txt"

    salvar_csv_ou_pasta(xml_path, propriedades, nome_saida_csv)

    gerar_txt = input("Deseja gerar um arquivo TXT, além do CSV? (s/n): ").strip().lower()
    if gerar_txt == "s":
        salvar_txt(resultados, propriedades, nome_saida_txt)
    else:
        print("Arquivo TXT não será gerado.")


if __name__ == "__main__":
    main()